import torch
import pandas as pd
import numpy as np
import torch.optim as optim
from sklearn.model_selection import train_test_split
import torch.nn as nn
import random
from models.mlp import BlackBoxModel  
from explainers.dce import DistributionalCounterfactualExplainer  

import time  
from openpyxl import Workbook 
from openpyxl.utils.dataframe import dataframe_to_rows  

pd.set_option('display.max_columns', None)

import sys, os
sys.path.append(os.path.abspath("."))

def set_seed(seed=42):
    os.environ["PYTHONHASHSEED"] = str(seed)
    random.seed(seed)          
    np.random.seed(seed)       
    torch.manual_seed(seed)    
    torch.cuda.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False

# ===================== Global Configuration =====================
seed_range = [40, 41, 42, 43, 44]  
output_excel = "cardio_DCE.xlsx"  


# ===================== 1. Define Target Column Name =====================
target_name = "cardio"  

# ===================== 2. Data Loading and Preprocessing =====================
seed = 42
set_seed(seed)
df_cardio = pd.read_csv('data/cardio/cardio.csv', delimiter=';')

# Target variable processing
df_X = df_cardio.iloc[:, 1:].copy()
df_X = df_X.drop(columns=[target_name]).copy()
df_y = df_cardio[target_name]

# Split into training and test sets (80% train, 20% test)
X_train, X_test, y_train, y_test = train_test_split(df_X, df_y, test_size=0.2, random_state=seed)

# Normalization (using mean and standard deviation of training set)
std = X_train.std()
mean = X_train.mean()
X_train = (X_train - mean) / std
X_test = (X_test - mean) / std

# Convert to PyTorch tensors
X_train_tensor = torch.FloatTensor(X_train.values)
y_train_tensor = torch.FloatTensor(y_train.values).view(-1, 1)
X_test_tensor = torch.FloatTensor(X_test.values)
y_test_tensor = torch.FloatTensor(y_test.values).view(-1, 1)

# Set device (GPU if available)
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# ===================== 3. Model Training =====================
# Reset random seed before model initialization
torch.manual_seed(seed)

# Initialize model, loss function, optimizer
model = BlackBoxModel(input_dim=X_train.shape[1]).to(device)
criterion = nn.MSELoss()
optimizer = optim.Adam(model.parameters(), lr=0.01)

# Training loop
num_epochs = 300
for epoch in range(num_epochs):
    model.train()
    # Forward pass
    outputs = model(X_train_tensor.to(device))
    loss = criterion(outputs, y_train_tensor.to(device))
    
    # Backward pass and optimization
    optimizer.zero_grad()
    loss.backward()
    optimizer.step()

# Evaluate model
model.eval()
with torch.no_grad():
    test_outputs = model(X_test_tensor.to(device))
    test_loss = criterion(test_outputs, y_test_tensor.to(device))
    
    # Convert to binary labels (0.5 as threshold)
    y_pred_tensor = (test_outputs > 0.5).float()
    correct_predictions = (y_pred_tensor == y_test_tensor.to(device)).float().sum()
    accuracy = correct_predictions / y_test_tensor.shape[0]

print(f"Test Accuracy: {accuracy.item() * 100:.2f}%")

# ===================== 4. Prepare Data for Explanation (Key: Keep Normalized) =====================
sample_num = 100
delta = 0.1
alpha = 0.05
N = 10
explain_columns = df_X.columns.tolist()
# explain_columns = ["height", "weight", "ap_hi", "ap_lo", "cholesterol", "gluc", "smoke", "alco", "active"]

# Filter samples where ap_hi > ap_lo
df_filtered = X_test[X_test['ap_hi'] > X_test['ap_lo']]

# Randomly select filtered samples
indice = df_filtered.sample(sample_num, random_state=seed).index

# Get explainer input data (ensure normalized)
df_explain = df_filtered.loc[indice].copy()  
y_true = y_test.loc[indice]

# Model predictions for original samples
with torch.no_grad():
    y_factual = model(torch.FloatTensor(df_explain.values).to(device))  # Model outputs for factual samples

# Generate counterfactual target y_target (Beta distribution)
y_target = torch.distributions.beta.Beta(0.1, 0.9).sample((sample_num,)).to(device)
# Adjust dimension to (sample_num, 1) to avoid dimension mismatch
y_target = y_target.unsqueeze(1)

# ===================== 5. Save Original Samples to be Explained (After Denormalization) =====================
df_explain_original = df_explain * std.values + mean.values  # Denormalize to original scale
df_explain_original[target_name] = y_true.values  # True labels
df_explain_original[f"{target_name}_target"] = y_target.cpu().detach().numpy()  # Counterfactual targets
df_explain_original[f"{target_name}_factual_pred"] = y_factual.cpu().detach().numpy()  # Model predictions for factual samples
df_explain_original.to_csv('cardio_explained_samples.csv', index=False)

# I. Original DCE
# ===================== 6. DCE Explainer Initialization and Optimization (Run on GPU) =====================
from explainers.dce import DistributionalCounterfactualExplainer

# New 1: Initialize Excel workbook (before DCE explainer initialization)
wb = Workbook()  # Create Excel workbook
wb.remove(wb.active)  # Remove default Sheet
# New 2: Record DCE start time
start_time = time.time()
# New 3: Initialize metrics list (fix undefined issue)
metrics_list = []  # Add this line

# ===================== New: Wrap with seed loop =====================
for current_seed in seed_range:  
    start_time = time.time()
    
    set_seed(current_seed)  


    explainer = DistributionalCounterfactualExplainer(
        model=model, 
        df_X=df_explain,  # Pass normalized data!!!
        explain_columns=explain_columns,
        y_target=y_target, 
        lr=1e-1, 
        n_proj=N,
        delta=delta)

    # Optimization process (automatically run on GPU since model/tensors are on GPU)
    set_seed(current_seed)
    explainer.optimize(U_1=0.5, U_2=0.3, l=0.2, r=1, max_iter=50, tau=1e3)

    # ===================== 7. Extract and Process Counterfactual Results (Move to CPU) =====================
    # Extract counterfactual X (best results) and counterfactual Y (model-predicted counterfactual outputs)
    X_s = explainer.best_X[:, explainer.explain_indices].clone().cpu()  # Move to CPU
    y_counterfactual = explainer.y.clone().cpu()  # Counterfactual Y from explainer (model predictions)

    # Denormalize to original scale
    X_s = X_s * std.values + mean.values

    # ===================== 8. Convert to DataFrame and Save =====================
    X_s_df = pd.DataFrame(X_s.detach().numpy(), columns=explain_columns)

    # Add counterfactual-related target columns
    # 1. Binarized counterfactual Y (model-predicted counterfactual labels)
    X_s_df[target_name] = (y_counterfactual.detach().numpy() > 0.5).astype(int)
    # 2. Counterfactual target y_target (original continuous values)
    X_s_df[f"{target_name}_target"] = y_target.cpu().detach().numpy()
    # 3. Model-predicted counterfactual continuous values
    X_s_df[f"{target_name}_counterfactual_pred"] = y_counterfactual.detach().numpy()

    # Save results
    # X_s_df.to_csv('cardio_counterfactual.csv', index=False)
    ws = wb.create_sheet(title=f"DCE_Seed_{current_seed}")  # Create Seed_42 Sheet
    for r in dataframe_to_rows(X_s_df, index=False, header=True):
        ws.append(r)


    # ===================== Calculate Wasserstein Distance (Simplified Implementation) =====================
    # Extract X_s (counterfactual features) and X_t (original features) on GPU
    X_s = explainer.best_X[:, explainer.explain_indices]  
    X_t = explainer.X_prime[:, explainer.explain_indices]  
    # Calculate X SWD
    x_wasserstein = np.sqrt(explainer.swd.distance(X_s, X_t, delta)[0].item())

    # Calculate Y WD
    y_s = explainer.best_y
    y_t = explainer.y_prime
    y_wasserstein = explainer.wd.distance(y_s, y_t, delta)[0].item()

    # ===================== New: Record Runtime + Metrics =====================
    run_time = round((time.time() - start_time)/60, 2)  # Convert to minutes, keep 2 decimals
    metrics_list.append({
        "seed": current_seed,  # Fixed seed=42
        "X_SWD": round(x_wasserstein, 6),
        "Y_WD": round(y_wasserstein, 6),
        "run_time_min": run_time
    })
    # ===================== Print Wasserstein Distance Results =====================
    print(f"X Sliced Wasserstein Distance (SWD): {x_wasserstein:.6f}")
    print(f"Y Wasserstein Distance (WD): {y_wasserstein:.6f}")


# ===================== New: Write Metrics_Summary Sheet =====================
df_metrics = pd.DataFrame(metrics_list)
ws_metrics = wb.create_sheet(title="DCE_Metrics_Summary")
for r in dataframe_to_rows(df_metrics, index=False, header=True):
    ws_metrics.append(r)

wb.save("cardio_DCE.xlsx") 

# II. Mean Constraint
# ===================== Core Configuration for Mean Constraint (Keep Your Logic) =====================
target_value = 75
normalized_value = (target_value - mean["weight"]) / std["weight"]

from explainers.dce_v2 import DCEWithConstraints  
from explainers.manager import ConstraintManager  

wb_constraint = Workbook()
wb_constraint.remove(wb_constraint.active)
metrics_constraint_list = []

seed_range = [40, 41, 42, 43, 44]
for current_seed in seed_range:
    start_time = time.time()
    
    # Constraint configuration
    config = [{"type": "mean", "bounds": {"weight": (normalized_value,1000 )}, "lambda": 100}]
    cm_75_100 = ConstraintManager(configs=config, feature_names=explain_columns)

    set_seed(current_seed)
    explainer_MeanConstraint = DCEWithConstraints(
        model=model, 
        df_X=df_explain, 
        explain_columns=explain_columns,
        y_target=y_target, 
        lr=1e-1, 
        n_proj=N,
        delta=delta
    )
    explainer_MeanConstraint.constraint_manager = cm_75_100

    # Optimization process
    set_seed(current_seed)
    explainer_MeanConstraint.optimize(  
        U_1=0.5, 
        U_2=0.3, 
        l=0.2, 
        r=1, 
        max_iter=50, 
        tau=1e3
    )

    # 1. Move GPU tensors to CPU to avoid errors when converting to NumPy
    best_X_MeanConstraint = explainer_MeanConstraint.best_X.clone().cpu()  
    best_y_MeanConstraint = explainer_MeanConstraint.best_y.clone().cpu()  

    # Restore original scale (denormalization, keep your logic)
    best_X_MeanConstraint = best_X_MeanConstraint * std.values + mean.values  

    # Convert counterfactual results to DataFrame (keep your logic)
    X_s_df_MeanConstraint = pd.DataFrame(best_X_MeanConstraint.cpu().detach().numpy(), columns=explain_columns)
    X_s_df_MeanConstraint[f"{target_name}"] = (best_y_MeanConstraint.cpu().detach().numpy() > 0.5).astype(int)
    X_s_df_MeanConstraint[f"{target_name}_target"] = y_target.cpu().detach().numpy()
    X_s_df_MeanConstraint[f"{target_name}_counterfactual_pred"] = best_y_MeanConstraint.detach().numpy()

    # ===================== Adaptation: Save independent Sheet + CSV for each seed =====================
    ws = wb_constraint.create_sheet(title=f"Mean_Seed_{current_seed}")
    for r in dataframe_to_rows(X_s_df_MeanConstraint, index=False, header=True):
        ws.append(r)

    # ===================== Calculate Wasserstein Distance (Keep Your Logic) =====================
    X_s = explainer_MeanConstraint.best_X[:, explainer_MeanConstraint.explain_indices]
    X_t = explainer_MeanConstraint.X_prime[:, explainer_MeanConstraint.explain_indices]
    x_wasserstein_mean = np.sqrt(explainer_MeanConstraint.swd.distance(X_s, X_t, delta)[0].item())
    y_s = explainer_MeanConstraint.best_y
    y_t = explainer_MeanConstraint.y_prime
    y_wasserstein_mean = explainer_MeanConstraint.wd.distance(y_s, y_t, delta)[0].item()

    # ===================== Record metrics for current seed =====================
    run_time = round((time.time() - start_time)/60, 2)
    metrics_constraint_list.append({
        "seed": current_seed,
        "X_SWD_mean": round(x_wasserstein_mean, 6),
        "Y_WD_mean": round(y_wasserstein_mean, 6),
        "run_time_min": run_time
    })

    # ===================== Print constraint results for current seed =====================
    print(f"X Sliced Wasserstein Distance (SWD) with Mean constraint: {x_wasserstein_mean:.6f}")
    print(f"Y Wasserstein Distance (WD) with Mean constraint: {y_wasserstein_mean:.6f}")

# ===================== Save Metrics_Summary for Constrained Version =====================
df_metrics_constraint = pd.DataFrame(metrics_constraint_list)
ws_metrics = wb_constraint.create_sheet(title="Mean_Metrics_Summary")
for r in dataframe_to_rows(df_metrics_constraint, index=False, header=True):
    ws_metrics.append(r)
wb_constraint.save("cardio_MeanConstraint.xlsx")

# III. Std Constraint
# ===================== Core Configuration for Std Constraint =====================
target_value = 10
normalized_value = float(target_value / std["ap_lo"])

from explainers.dce_v2 import DCEWithConstraints  
from explainers.manager import ConstraintManager  

wb_std_constraint = Workbook()
wb_std_constraint.remove(wb_std_constraint.active)
# Initialize metrics list for constrained version
metrics_std_constraint_list = []

seed_range = [40, 41, 42, 43, 44]
for current_seed in seed_range:
    # Record runtime for current seed
    start_time = time.time()
    
    # 1. Constraint configuration: upper bound constraint on std of Credit amount (normalized space)
    configs_std_10_100 = [
        {"type": "std", "bounds": {"ap_lo": normalized_value}, "lambda": 100}
    ]
    cm_std_10_100 = ConstraintManager(configs=configs_std_10_100, feature_names=explain_columns)
    set_seed(current_seed)

    explainer_StdConstraint = DCEWithConstraints(
        model=model, 
        df_X=df_explain, 
        explain_columns=explain_columns,
        y_target=y_target, 
        lr=1e-1, 
        n_proj=N,
        delta=delta,
        constraint_manager=cm_std_10_100  
    )

    # Optimization process
    set_seed(current_seed)
    explainer_StdConstraint.optimize(  
        U_1=0.5, 
        U_2=0.3, 
        l=0.2, 
        r=1, 
        max_iter=50, 
        tau=1e3
    )

    # 1. Move GPU tensors to CPU to avoid errors when converting to NumPy
    best_X_StdConstraint = explainer_StdConstraint.best_X.clone().cpu()  
    best_y_StdConstraint = explainer_StdConstraint.best_y.clone().cpu()  


    # Restore original scale (denormalization)
    best_X_StdConstraint = best_X_StdConstraint * std.values + mean.values  # Restore X_s to original scale

    # Save counterfactual results as DataFrame with the same format as original data
    X_s_df_StdConstraint = pd.DataFrame(best_X_StdConstraint.cpu().detach().numpy(), columns=explain_columns)

    # Binarize cardio using threshold 0.5
    X_s_df_StdConstraint[f"{target_name}"] = (best_y_StdConstraint.cpu().detach().numpy() > 0.5).astype(int)  # Convert y_target to 0 or 1
    X_s_df_StdConstraint[f"{target_name}_target"] = y_target.cpu().detach().numpy()  # Add y_target column
    X_s_df_StdConstraint[f"{target_name}_counterfactual_pred"] =best_y_StdConstraint.detach().numpy()

    # Save to Excel Sheet (with seed identifier)
    ws = wb_std_constraint.create_sheet(title=f"Std_Seed_{current_seed}")
    for r in dataframe_to_rows(X_s_df_StdConstraint, index=False, header=True):
        ws.append(r)

    # ===================== Calculate Wasserstein Distance =====================
    # Extract X_s (counterfactual features) and X_t (original features) on GPU
    X_s = explainer_StdConstraint.best_X[:, explainer_StdConstraint.explain_indices]  
    X_t = explainer_StdConstraint.X_prime[:, explainer_StdConstraint.explain_indices]  
    # Calculate X SWD
    x_wasserstein_std = np.sqrt(explainer_StdConstraint.swd.distance(X_s, X_t, delta)[0].item())

    # Calculate Y WD
    y_s = explainer_StdConstraint.best_y
    y_t = explainer_StdConstraint.y_prime
    y_wasserstein_std = explainer_StdConstraint.wd.distance(y_s, y_t, delta)[0].item()

    # ===================== Record metrics for current seed =====================
    run_time = round((time.time() - start_time)/60, 2)
    metrics_std_constraint_list.append({
        "seed": current_seed,
        "X_SWD_std": round(x_wasserstein_std, 6),
        "Y_WD_std": round(y_wasserstein_std, 6),
        "run_time_min": run_time
    })

    # ===================== Print Wasserstein Distance Results =====================
    print(f"X Sliced Wasserstein Distance (SWD) with std constraint: {x_wasserstein_std:.6f}")
    print(f"Y Wasserstein Distance (WD) with std constraint: {y_wasserstein_std:.6f}")

# ===================== Save Metrics_Summary for Constrained Version =====================
df_metrics_std_constraint = pd.DataFrame(metrics_std_constraint_list)
ws_std_metrics = wb_std_constraint.create_sheet(title="Std_Metrics_Summary")
for r in dataframe_to_rows(df_metrics_std_constraint, index=False, header=True):
    ws_std_metrics.append(r)
# Save constrained Excel
wb_std_constraint.save("cardio_StdConstraint.xlsx")

# IV. LSC Constraint
# ===================== Core Configuration for LSC Constraint =====================
target_value = 0.8

from explainers.dce_v2 import DCEWithConstraints  
from explainers.manager import ConstraintManager  

# Initialize Excel workbook (save constraint results for all seeds)
wb_lsc_constraint = Workbook()
wb_lsc_constraint.remove(wb_lsc_constraint.active)
# Initialize metrics list for constrained version
metrics_lsc_constraint_list = []

seed_range = [40, 41, 42, 43, 44]
for current_seed in seed_range:
    start_time = time.time()
    
    # 1. Constraint configuration: upper bound constraint on std of Credit amount (normalized space)
    configs_hinge = [
        {
            "type": "lsc",
            "relation": {
                "cholesterol": (["gluc"], [0.4557], 0.0)
            },
            "lambda": 10,
            "mode": "hinge",
            "tolerance":  target_value  
        }
    ]

    # Ensure order matches df_explain.columns
    feature_names = list(df_explain.columns)
    constraint_manager_hinge = ConstraintManager(configs=configs_hinge, feature_names=feature_names)

    set_seed(current_seed)
    explainer_lsc_hinge = DCEWithConstraints(
        model=model, 
        df_X=df_explain, 
        explain_columns=explain_columns,
        y_target=y_target, 
        lr=1e-1, 
        n_proj=N,
        delta=delta,
        constraint_manager=constraint_manager_hinge
    )

    # Optimization process
    set_seed(current_seed)
    explainer_lsc_hinge.optimize(
        U_1=0.5, 
        U_2=0.3, 
        l=0.2, 
        r=1, 
        max_iter=50, 
        tau=10
    )

    # 1. Move GPU tensors to CPU to avoid errors when converting to NumPy
    best_X_lsc_hinge = explainer_lsc_hinge.best_X.clone().cpu()  
    best_y_lsc_hinge = explainer_lsc_hinge.best_y.clone().cpu()  


    # Restore original scale (denormalization)
    best_X_lsc_hinge = best_X_lsc_hinge * std.values + mean.values  # Restore X_s to original scale

    # Save counterfactual results as DataFrame with the same format as original data
    X_s_df_lsc_hinge = pd.DataFrame(best_X_lsc_hinge.cpu().detach().numpy(), columns=explain_columns)

    # Binarize cardio using threshold 0.5
    X_s_df_lsc_hinge[f"{target_name}"] = (best_y_lsc_hinge.cpu().detach().numpy() > 0.5).astype(int)  # Convert y_target to 0 or 1
    X_s_df_lsc_hinge[f"{target_name}_target"] = y_target.cpu().detach().numpy()  # Add y_target column
    X_s_df_lsc_hinge[f"{target_name}_counterfactual_pred"] =best_y_lsc_hinge.detach().numpy()


    # Save to Excel Sheet (with seed identifier)
    ws = wb_lsc_constraint.create_sheet(title=f"LSC_Seed_{current_seed}")
    for r in dataframe_to_rows(X_s_df_lsc_hinge, index=False, header=True):
        ws.append(r)

    # ===================== Calculate Wasserstein Distance =====================
    # Extract X_s (counterfactual features) and X_t (original features) on GPU
    X_s = explainer_lsc_hinge.best_X[:, explainer_lsc_hinge.explain_indices]  # GPU tensor, no longer overwrite CPU variable
    X_t = explainer_lsc_hinge.X_prime[:, explainer_lsc_hinge.explain_indices]  # Take columns matching X_s for dimension consistency
    # Calculate X SWD with your specified implementation
    x_wasserstein_lsc = np.sqrt(explainer_lsc_hinge.swd.distance(X_s, X_t, delta)[0].item())

    # Calculate Y WD
    y_s = explainer_lsc_hinge.best_y
    y_t = explainer_lsc_hinge.y_prime
    y_wasserstein_lsc = explainer_lsc_hinge.wd.distance(y_s, y_t, delta)[0].item()

    # ===================== Record metrics for current seed =====================
    run_time = round((time.time() - start_time)/60, 2)
    metrics_lsc_constraint_list.append({
        "seed": current_seed,
        "X_SWD_lsc": round(x_wasserstein_lsc, 6),
        "Y_WD_lsc": round(y_wasserstein_lsc, 6),
        "run_time_min": run_time
    })

    # ===================== New: Print Wasserstein Distance Results =====================
    print(f"X Sliced Wasserstein Distance (SWD) with LSC constraint: {x_wasserstein_lsc:.6f}")
    print(f"Y Wasserstein Distance (WD) with LSC constraint: {y_wasserstein_lsc:.6f}")

# ===================== Save Metrics_Summary for Constrained Version =====================
df_metrics_lsc_constraint = pd.DataFrame(metrics_lsc_constraint_list)
ws_lsc_metrics = wb_lsc_constraint.create_sheet(title="LSC_Metrics_Summary")
for r in dataframe_to_rows(df_metrics_lsc_constraint, index=False, header=True):
    ws_lsc_metrics.append(r)
# Save constrained Excel
wb_lsc_constraint.save("cardio_LSCConstraint.xlsx")

# V. FSD Constraint
# ===================== Core Configuration for FSD Constraint =====================
from explainers.dce_v2 import DCEWithConstraints
from explainers.manager import ConstraintManager
from explainers.constraints.FSDConstraint import FSDConstraintTorch

# Initialize Excel workbook (save constraint results for all seeds)
wb_fsd_constraint = Workbook()
wb_fsd_constraint.remove(wb_fsd_constraint.active)
# Initialize metrics list for constrained version
metrics_fsd_constraint_list = []

seed_range = [40, 41, 42, 43, 44]
for current_seed in seed_range:
    start_time = time.time()
    
    # 1. Constraint configuration
    configs = [
        {
            "type": "fsd",
            "lambda": 1,   # Penalty coefficient
            "dir_map": {"weight": -1},   # Duration must shift left
            "M": 100,      # Number of samples for FSD calculation
            "sample_mode": "quantile"
        }
    ]

    # Ensure order matches df_explain.columns
    feature_names = list(df_explain.columns)
    constraint_manager_fsd = ConstraintManager(configs=configs, feature_names=feature_names)

    set_seed(current_seed)
    explainer_fsd = DCEWithConstraints(
        model=model, 
        df_X=df_explain, 
        explain_columns=explain_columns,
        y_target=y_target, 
        lr=1e-1, 
        n_proj=N,
        delta=delta,
        constraint_manager=constraint_manager_fsd
    )

    # Optimization process
    set_seed(current_seed)
    explainer_fsd.optimize(
        U_1=0.5, 
        U_2=0.3, 
        l=0.2, 
        r=1, 
        max_iter=50, 
        tau=10
    )

    # 1. Move GPU tensors to CPU to avoid errors when converting to NumPy
    best_X_fsd = explainer_fsd.best_X.clone().cpu()  
    best_y_fsd = explainer_fsd.best_y.clone().cpu()  


    # Restore original scale (denormalization)
    best_X_fsd = best_X_fsd * std.values + mean.values  # Restore X_s to original scale

    # Save counterfactual results as DataFrame with the same format as original data
    X_s_df_fsd = pd.DataFrame(best_X_fsd.cpu().detach().numpy(), columns=explain_columns)

    # Binarize cardio using threshold 0.5
    X_s_df_fsd[f"{target_name}"] = (best_y_fsd.cpu().detach().numpy() > 0.5).astype(int)  # Convert y_target to 0 or 1
    X_s_df_fsd[f"{target_name}_target"] = y_target.cpu().detach().numpy()  # Add y_target column
    X_s_df_fsd[f"{target_name}_counterfactual_pred"] =best_y_fsd.detach().numpy()

    # Save to Excel Sheet (with seed identifier)
    ws = wb_fsd_constraint.create_sheet(title=f"FSD_Seed_{current_seed}")
    for r in dataframe_to_rows(X_s_df_fsd, index=False, header=True):
        ws.append(r)

    # ===================== Calculate Wasserstein Distance =====================
    # Extract X_s (counterfactual features) and X_t (original features) on GPU
    X_s = explainer_fsd.best_X[:, explainer_fsd.explain_indices]  # GPU tensor, no longer overwrite CPU variable
    X_t = explainer_fsd.X_prime[:, explainer_fsd.explain_indices]  # Take columns matching X_s for dimension consistency
    # Calculate X SWD with your specified implementation
    x_wasserstein_fsd = np.sqrt(explainer_fsd.swd.distance(X_s, X_t, delta)[0].item())

    # Calculate Y WD
    y_s = explainer_fsd.best_y
    y_t = explainer_fsd.y_prime
    y_wasserstein_fsd = explainer_fsd.wd.distance(y_s, y_t, delta)[0].item()

    # ===================== Record metrics for current seed =====================
    run_time = round((time.time() - start_time)/60, 2)
    metrics_fsd_constraint_list.append({
        "seed": current_seed,
        "X_SWD_fsd": round(x_wasserstein_fsd, 6),
        "Y_WD_fsd": round(y_wasserstein_fsd, 6),
        "run_time_min": run_time
    })

    # ===================== Print Wasserstein Distance Results =====================
    print(f"X Sliced Wasserstein Distance (SWD) with FSD constraint: {x_wasserstein_fsd:.6f}")
    print(f"Y Wasserstein Distance (WD) with FSD constraint: {y_wasserstein_fsd:.6f}")

# ===================== Save Metrics_Summary for Constrained Version =====================
df_metrics_fsd_constraint = pd.DataFrame(metrics_fsd_constraint_list)
ws_fsd_metrics = wb_fsd_constraint.create_sheet(title="FSD_Metrics_Summary")
for r in dataframe_to_rows(df_metrics_fsd_constraint, index=False, header=True):
    ws_fsd_metrics.append(r)
# Save constrained Excel
wb_fsd_constraint.save("cardio_FSDConstraint.xlsx")