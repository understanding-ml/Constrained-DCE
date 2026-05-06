import torch
import pandas as pd
import numpy as np
import torch.optim as optim
from sklearn.model_selection import train_test_split
import torch.nn as nn
import random
from models.mlp import BlackBoxModel  
from explainers.dce import DistributionalCounterfactualExplainer  
from sklearn.preprocessing import LabelEncoder

import time  
from openpyxl import Workbook  
from openpyxl.utils.dataframe import dataframe_to_rows  

pd.set_option('display.max_columns', None)

import sys, os
sys.path.append(os.path.abspath("."))

def set_seed(seed=42):
    os.environ["PYTHONHASHSEED"] = str(seed)
    random.seed(seed)          
    np.random.seed(seed)       
    torch.manual_seed(seed)    
    torch.cuda.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False

# ===================== 1. Define Missing Target Column Name =====================
target_name = "Response"  # Define target_name to match the target column in the dataset

# ===================== 2. Data Loading and Preprocessing =====================
seed = 42
set_seed(seed)
df = pd.read_excel('data/Marketing Campaign/marketing_campaign.xlsx')
# Drop Z_CostContact and Z_Revenue columns
df = df.drop(columns=['Z_CostContact', 'Z_Revenue'])

label_encoder = LabelEncoder()
label_mappings = {}

# Sequential mapping
education_mapping = {
    'PhD': 3,
    'Master': 2,
    '2n Cycle': 2,
    'Graduation': 1,
    'Basic': 0
}
marital_mapping = {
    'Married': 1,
    'Together': 1,
    'Single': 0,
    'Divorced': 0,
    'Alone': 0,
    'YOLO': 0,
    'Absurd': 0
}

# Iterate through each column for mapping
for column in df.columns:
    if column != target_name and df[column].dtype == 'object':  # Exclude target variable
        if column == "Education":
            df[column] = df[column].map(education_mapping)  # Map using education_mapping
        elif column == "Marital_Status":
            df[column] = df[column].map(marital_mapping)  # Map using marital_mapping
        else:
            # Encode other categorical variables with LabelEncoder
            df[column] = label_encoder.fit_transform(df[column].astype(str))

# Key: Handle NaN only at the final step
for column in df.columns:
    if df[column].isna().any():
        median_val = df[column].median()
        df[column] = df[column].fillna(median_val).astype(int)

# Target variable processing
df_X = df.iloc[:, 1:].copy()
# Drop target variable 'Response' column
df_X = df_X.drop(columns=['Response'])
df_y = df[target_name]

# Split into training and test sets (80% train, 20% test)
X_train, X_test, y_train, y_test = train_test_split(df_X, df_y, test_size=0.2, random_state=seed)

# Normalization (using mean and standard deviation of training set)
std = X_train.std()
mean = X_train.mean()
X_train = (X_train - mean) / std
X_test = (X_test - mean) / std

# Convert to PyTorch tensors
X_train_tensor = torch.FloatTensor(X_train.values)
y_train_tensor = torch.FloatTensor(y_train.values).view(-1, 1)
X_test_tensor = torch.FloatTensor(X_test.values)
y_test_tensor = torch.FloatTensor(y_test.values).view(-1, 1)

# Set device (GPU if available)
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# ===================== 3. Model Training =====================
# Force reset random seed before model initialization
torch.manual_seed(seed)

# Initialize model, loss function, optimizer
model = BlackBoxModel(input_dim=X_train.shape[1]).to(device)
criterion = nn.MSELoss()
optimizer = optim.Adam(model.parameters(), lr=0.01)

# Training loop
num_epochs = 300
for epoch in range(num_epochs):
    model.train()
    # Forward pass
    outputs = model(X_train_tensor.to(device))
    loss = criterion(outputs, y_train_tensor.to(device))
    
    # Backward pass and optimization
    optimizer.zero_grad()
    loss.backward()
    optimizer.step()

# Evaluate model
model.eval()
with torch.no_grad():
    test_outputs = model(X_test_tensor.to(device))
    test_loss = criterion(test_outputs, y_test_tensor.to(device))
    
    # Convert to binary labels (0.5 as threshold)
    y_pred_tensor = (test_outputs > 0.5).float()
    correct_predictions = (y_pred_tensor == y_test_tensor.to(device)).float().sum()
    accuracy = correct_predictions / y_test_tensor.shape[0]

print(f"Test Accuracy: {accuracy.item() * 100:.2f}%")

# ===================== 4. Prepare Data for Explanation (Key: Keep Normalized) =====================
sample_num = 100
delta = 0.1
alpha = 0.05
N = 10
explain_columns = df_X.columns.tolist()

# Randomly select test set samples
indice = X_test.sample(sample_num, random_state=seed).index
df_explain = X_test.loc[indice].copy()  # Input to explainer must be normalized data!!!
y_true = y_test.loc[indice]

# Model predictions for original samples
with torch.no_grad():
    y_factual = model(torch.FloatTensor(df_explain.values).to(device))  # Model outputs for factual samples

# Generate counterfactual target y_target (Beta distribution)
y_target = torch.distributions.beta.Beta(0.5, 0.5).sample((sample_num,)).to(device)
# Adjust dimension to (sample_num, 1) to avoid dimension mismatch
y_target = y_target.unsqueeze(1)

# ===================== 5. Save Original Samples to be Explained (After Denormalization) =====================
df_explain_original = df_explain * std.values + mean.values  # Denormalize to restore original scale
df_explain_original[target_name] = y_true.values  # True labels
df_explain_original[f"{target_name}_target"] = y_target.cpu().detach().numpy()  # Counterfactual targets
df_explain_original[f"{target_name}_factual_pred"] = y_factual.cpu().detach().numpy()  # Model predictions for factual samples
df_explain_original.to_csv('market_explained_samples.csv', index=False)


# ===================== I. Original DCE =====================
from explainers.dce import DistributionalCounterfactualExplainer
# Initialize Excel workbook
wb_dce = Workbook()
wb_dce.remove(wb_dce.active)
metrics_dce_list = []
seed_range = [40,41,42,43,44]

for current_seed in seed_range:
    start_time = time.time()
    set_seed(current_seed)
    explainer = DistributionalCounterfactualExplainer(
        model=model, 
        df_X=df_explain,
        explain_columns=explain_columns,
        y_target=y_target, 
        lr=1e-1, 
        n_proj=N,
        delta=delta)

    explainer.optimize(U_1=0.5, U_2=0.3, l=0.2, r=1, max_iter=50, tau=1e3)

    # Extract results
    X_s = explainer.best_X[:, explainer.explain_indices].clone().cpu()
    y_counterfactual = explainer.y.clone().cpu()
    X_s = X_s * std.values + mean.values

    # Convert to DataFrame
    X_s_df = pd.DataFrame(X_s.detach().numpy(), columns=explain_columns)
    X_s_df[target_name] = (y_counterfactual.detach().numpy() > 0.5).astype(int)
    X_s_df[f"{target_name}_target"] = y_target.cpu().detach().numpy()
    X_s_df[f"{target_name}_counterfactual_pred"] = y_counterfactual.detach().numpy()
    # X_s_df.to_csv(f'market_counterfactual_seed_{current_seed}.csv', index=False)

    # Save to Excel Sheet
    ws = wb_dce.create_sheet(title=f"DCE_Seed_{current_seed}")
    for r in dataframe_to_rows(X_s_df, index=False, header=True):
        ws.append(r)

    # Calculate Wasserstein distance
    X_s_gpu = explainer.best_X[:, explainer.explain_indices]
    X_t = explainer.X_prime[:, explainer.explain_indices]
    x_wasserstein = np.sqrt(explainer.swd.distance(X_s_gpu, X_t, delta)[0].item())

    y_s = explainer.best_y
    y_t = explainer.y_prime
    y_wasserstein = explainer.wd.distance(y_s, y_t, delta)[0].item()

    # Record metrics
    run_time = round((time.time() - start_time)/60, 2)
    metrics_dce_list.append({
        "seed": current_seed,
        "X_SWD": round(x_wasserstein, 6),
        "Y_WD": round(y_wasserstein, 6),
        "run_time_min": run_time
    })

    print(f"X Sliced Wasserstein Distance (SWD): {x_wasserstein:.6f}")
    print(f"Y Wasserstein Distance (WD): {y_wasserstein:.6f}")

# Save metrics summary sheet
df_metrics_dce = pd.DataFrame(metrics_dce_list)
ws_metrics = wb_dce.create_sheet(title="DCE_Metrics_Summary")
for r in dataframe_to_rows(df_metrics_dce, index=False, header=True):
    ws_metrics.append(r)
wb_dce.save("market_DCE.xlsx")

# ===================== II. Mean Constraint =====================
target_value = 54
normalized_value = (target_value - mean["Recency"]) / std["Recency"]

from explainers.dce_v2 import DCEWithConstraints  
from explainers.manager import ConstraintManager  
config = [{"type": "mean", "bounds": {"Recency": (normalized_value,1000 )}, "lambda": 100}]
cm_54_100 = ConstraintManager(configs=config, feature_names=explain_columns)

# Initialize Excel workbook
wb_mean = Workbook()
wb_mean.remove(wb_mean.active)
metrics_mean_list = []
seed_range = [40,41,42,43,44]

for current_seed in seed_range:
    start_time = time.time()
    set_seed(current_seed)
    explainer_MeanConstraint = DCEWithConstraints(
        model=model, 
        df_X=df_explain, 
        explain_columns=explain_columns,
        y_target=y_target, 
        lr=1e-1, 
        n_proj=N,
        delta=delta
    )
    explainer_MeanConstraint.constraint_manager = cm_54_100
    explainer_MeanConstraint.optimize(U_1=0.5, U_2=0.3, l=0.2, r=1, max_iter=50, tau=1e3)

    # Extract results
    best_X_MeanConstraint = explainer_MeanConstraint.best_X.clone().clone().cpu()  
    best_y_MeanConstraint = explainer_MeanConstraint.best_y.clone().clone().cpu()  
    best_X_MeanConstraint = best_X_MeanConstraint * std.values + mean.values

    X_s_df_MeanConstraint = pd.DataFrame(best_X_MeanConstraint.cpu().detach().numpy(), columns=explain_columns)
    X_s_df_MeanConstraint[f"{target_name}"] = (best_y_MeanConstraint.cpu().detach().numpy() > 0.5).astype(int)
    X_s_df_MeanConstraint[f"{target_name}_target"] = y_target.cpu().detach().numpy()
    X_s_df_MeanConstraint[f"{target_name}_counterfactual_pred"] =best_y_MeanConstraint.detach().numpy()
    # X_s_df_MeanConstraint.to_csv(f'market_MeanConstraint_54_100_seed_{current_seed}.csv', index=False)

    # Save to Excel
    ws = wb_mean.create_sheet(title=f"Mean_Seed_{current_seed}")
    for r in dataframe_to_rows(X_s_df_MeanConstraint, index=False, header=True):
        ws.append(r)

    # Calculate distance
    X_s_gpu = explainer_MeanConstraint.best_X[:, explainer_MeanConstraint.explain_indices]
    X_t = explainer_MeanConstraint.X_prime[:, explainer_MeanConstraint.explain_indices]
    x_wasserstein_mean = np.sqrt(explainer_MeanConstraint.swd.distance(X_s_gpu, X_t, delta)[0].item())

    y_s = explainer_MeanConstraint.best_y
    y_t = explainer_MeanConstraint.y_prime
    y_wasserstein_mean = explainer_MeanConstraint.wd.distance(y_s, y_t, delta)[0].item()

    # Record metrics
    run_time = round((time.time() - start_time)/60, 2)
    metrics_mean_list.append({
        "seed": current_seed,
        "X_SWD_mean": round(x_wasserstein_mean, 6),
        "Y_WD_mean": round(y_wasserstein_mean, 6),
        "run_time_min": run_time
    })

    print(f"X Sliced Wasserstein Distance (SWD) with Mean constraint: {x_wasserstein_mean:.6f}")
    print(f"Y Wasserstein Distance (WD) with Mean constraint: {y_wasserstein_mean:.6f}")

# Metrics summary
df_metrics_mean = pd.DataFrame(metrics_mean_list)
ws_metrics = wb_mean.create_sheet(title="Mean_Metrics_Summary")
for r in dataframe_to_rows(df_metrics_mean, index=False, header=True):
    ws_metrics.append(r)
wb_mean.save("market_MeanConstraint.xlsx")

# ===================== III. Std Constraint =====================
target_value = 19000
normalized_value = float(target_value / std["Income"])

from explainers.dce_v2 import DCEWithConstraints  
from explainers.manager import ConstraintManager  
configs_std_19000_100 = [
    {"type": "std", "bounds": {"Income": normalized_value}, "lambda": 100}
]
cm_std_19000_100 = ConstraintManager(configs=configs_std_19000_100, feature_names=explain_columns)

# Initialize Excel workbook
wb_std = Workbook()
wb_std.remove(wb_std.active)
metrics_std_list = []
seed_range = [40,41,42,43,44]

for current_seed in seed_range:
    start_time = time.time()
    set_seed(current_seed)
    explainer_StdConstraint = DCEWithConstraints(
        model=model, 
        df_X=df_explain, 
        explain_columns=explain_columns,
        y_target=y_target, 
        lr=1e-1, 
        n_proj=N,
        delta=delta,
        constraint_manager=cm_std_19000_100
    )
    explainer_StdConstraint.optimize(U_1=0.5, U_2=0.3, l=0.2, r=1, max_iter=50, tau=1e3)

    # Extract results
    best_X_StdConstraint = explainer_StdConstraint.best_X.clone().clone().cpu()  
    best_y_StdConstraint = explainer_StdConstraint.best_y.clone().clone().cpu()  
    best_X_StdConstraint = best_X_StdConstraint * std.values + mean.values

    X_s_df_StdConstraint = pd.DataFrame(best_X_StdConstraint.cpu().detach().numpy(), columns=explain_columns)
    X_s_df_StdConstraint[f"{target_name}"] = (best_y_StdConstraint.cpu().detach().numpy() > 0.5).astype(int)
    X_s_df_StdConstraint[f"{target_name}_target"] = y_target.cpu().detach().numpy()
    X_s_df_StdConstraint[f"{target_name}_counterfactual_pred"] =best_y_StdConstraint.detach().numpy()
    # X_s_df_StdConstraint.to_csv(f'market_StdConstraint_19000_100_seed_{current_seed}.csv', index=False)

    # Save to Excel
    ws = wb_std.create_sheet(title=f"Std_Seed_{current_seed}")
    for r in dataframe_to_rows(X_s_df_StdConstraint, index=False, header=True):
        ws.append(r)

    # Calculate distance
    X_s_gpu = explainer_StdConstraint.best_X[:, explainer_StdConstraint.explain_indices]
    X_t = explainer_StdConstraint.X_prime[:, explainer_StdConstraint.explain_indices]
    x_wasserstein_std = np.sqrt(explainer_StdConstraint.swd.distance(X_s_gpu, X_t, delta)[0].item())

    y_s = explainer_StdConstraint.best_y
    y_t = explainer_StdConstraint.y_prime
    y_wasserstein_std = explainer_StdConstraint.wd.distance(y_s, y_t, delta)[0].item()

    # Record metrics
    run_time = round((time.time() - start_time)/60, 2)
    metrics_std_list.append({
        "seed": current_seed,
        "X_SWD_std": round(x_wasserstein_std, 6),
        "Y_WD_std": round(y_wasserstein_std, 6),
        "run_time_min": run_time
    })

    print(f"X Sliced Wasserstein Distance (SWD) with std constraint: {x_wasserstein_std:.6f}")
    print(f"Y Wasserstein Distance (WD) with std constraint: {y_wasserstein_std:.6f}")

# Metrics summary
df_metrics_std = pd.DataFrame(metrics_std_list)
ws_metrics = wb_std.create_sheet(title="Std_Metrics_Summary")
for r in dataframe_to_rows(df_metrics_std, index=False, header=True):
    ws_metrics.append(r)
wb_std.save("market_StdConstraint.xlsx")

# ===================== IV. LSC Constraint =====================
target_value = 0.66

from explainers.dce_v2 import DCEWithConstraints  
from explainers.manager import ConstraintManager  
configs_hinge = [
    {
        "type": "lsc",
        "relation": {
            "MntFruits": (["MntSweetProducts"], [0.5785], 0.0)
        },
        "lambda": 1,
        "mode": "hinge",
        "tolerance":  target_value  
    }
]

feature_names = list(df_explain.columns)
constraint_manager_hinge = ConstraintManager(configs=configs_hinge, feature_names=feature_names)

# Initialize Excel workbook
wb_lsc = Workbook()
wb_lsc.remove(wb_lsc.active)
metrics_lsc_list = []
seed_range = [40,41,42,43,44]

for current_seed in seed_range:
    start_time = time.time()
    set_seed(current_seed)
    explainer_lsc_hinge = DCEWithConstraints(
        model=model, 
        df_X=df_explain, 
        explain_columns=explain_columns,
        y_target=y_target, 
        lr=10, 
        n_proj=N,
        delta=delta,
        constraint_manager=constraint_manager_hinge
    )
    explainer_lsc_hinge.optimize(U_1=0.5, U_2=0.3, l=0.2, r=1, max_iter=50, tau=10)

    # Extract results
    best_X_lsc_hinge = explainer_lsc_hinge.best_X.clone().clone().cpu()  
    best_y_lsc_hinge = explainer_lsc_hinge.best_y.clone().clone().cpu()  
    best_X_lsc_hinge = best_X_lsc_hinge * std.values + mean.values

    X_s_df_lsc_hinge = pd.DataFrame(best_X_lsc_hinge.cpu().detach().numpy(), columns=explain_columns)
    X_s_df_lsc_hinge[f"{target_name}"] = (best_y_lsc_hinge.cpu().detach().numpy() > 0.5).astype(int)
    X_s_df_lsc_hinge[f"{target_name}_target"] = y_target.cpu().detach().numpy()
    X_s_df_lsc_hinge[f"{target_name}_counterfactual_pred"] =best_y_lsc_hinge.detach().numpy()
    # X_s_df_lsc_hinge.to_csv(f'market_lsc_hinge_seed_{current_seed}.csv', index=False)

    # Save to Excel
    ws = wb_lsc.create_sheet(title=f"LSC_Seed_{current_seed}")
    for r in dataframe_to_rows(X_s_df_lsc_hinge, index=False, header=True):
        ws.append(r)

    # Calculate distance
    X_s_gpu = explainer_lsc_hinge.best_X[:, explainer_lsc_hinge.explain_indices]
    X_t = explainer_lsc_hinge.X_prime[:, explainer_lsc_hinge.explain_indices]
    x_wasserstein_lsc = np.sqrt(explainer_lsc_hinge.swd.distance(X_s_gpu, X_t, delta)[0].item())

    y_s = explainer_lsc_hinge.best_y
    y_t = explainer_lsc_hinge.y_prime
    y_wasserstein_lsc = explainer_lsc_hinge.wd.distance(y_s, y_t, delta)[0].item()

    # Record metrics
    run_time = round((time.time() - start_time)/60, 2)
    metrics_lsc_list.append({
        "seed": current_seed,
        "X_SWD_lsc": round(x_wasserstein_lsc, 6),
        "Y_WD_lsc": round(y_wasserstein_lsc, 6),
        "run_time_min": run_time
    })

    print(f"X Sliced Wasserstein Distance (SWD) with LSC constraint: {x_wasserstein_lsc:.6f}")
    print(f"Y Wasserstein Distance (WD) with LSC constraint: {y_wasserstein_lsc:.6f}")

# Metrics summary
df_metrics_lsc = pd.DataFrame(metrics_lsc_list)
ws_metrics = wb_lsc.create_sheet(title="LSC_Metrics_Summary")
for r in dataframe_to_rows(df_metrics_lsc, index=False, header=True):
    ws_metrics.append(r)
wb_lsc.save("market_LSCConstraint.xlsx")

# ===================== V. FSD Constraint =====================
from explainers.dce_v2 import DCEWithConstraints
from explainers.manager import ConstraintManager
from explainers.constraints.FSDConstraint import FSDConstraintTorch

configs = [
    {
        "type": "fsd",
        "lambda": 0.004,
        "dir_map": {"NumDealsPurchases": 1},
        "M": 100,        
        "sample_mode": "quantile"
    }
]

feature_names = list(df_explain.columns)
constraint_manager_fsd = ConstraintManager(configs=configs, feature_names=feature_names)

# Initialize Excel workbook
wb_fsd = Workbook()
wb_fsd.remove(wb_fsd.active)
metrics_fsd_list = []
seed_range = [40,41,42,43,44]

for current_seed in seed_range:
    start_time = time.time()
    set_seed(current_seed)
    explainer_fsd = DCEWithConstraints(
        model=model, 
        df_X=df_explain, 
        explain_columns=explain_columns,
        y_target=y_target, 
        lr=10, 
        n_proj=N,
        delta=delta,
        constraint_manager=constraint_manager_fsd
    )
    explainer_fsd.optimize(U_1=0.5, U_2=0.3, l=0.2, r=1, max_iter=50, tau=10)

    # Extract results
    best_X_fsd = explainer_fsd.best_X.clone().clone().cpu()  
    best_y_fsd = explainer_fsd.best_y.clone().clone().cpu()  
    best_X_fsd = best_X_fsd * std.values + mean.values

    X_s_df_fsd = pd.DataFrame(best_X_fsd.cpu().detach().numpy(), columns=explain_columns)
    X_s_df_fsd[f"{target_name}"] = (best_y_fsd.cpu().detach().numpy() > 0.5).astype(int)
    X_s_df_fsd[f"{target_name}_target"] = y_target.cpu().detach().numpy()
    X_s_df_fsd[f"{target_name}_counterfactual_pred"] =best_y_fsd.detach().numpy()
    # X_s_df_fsd.to_csv(f'market_fsd_seed_{current_seed}.csv', index=False)

    # Save to Excel
    ws = wb_fsd.create_sheet(title=f"FSD_Seed_{current_seed}")
    for r in dataframe_to_rows(X_s_df_fsd, index=False, header=True):
        ws.append(r)

    # Calculate distance
    X_s_gpu = explainer_fsd.best_X[:, explainer_fsd.explain_indices]
    X_t = explainer_fsd.X_prime[:, explainer_fsd.explain_indices]
    x_wasserstein_fsd = np.sqrt(explainer_fsd.swd.distance(X_s_gpu, X_t, delta)[0].item())

    y_s = explainer_fsd.best_y
    y_t = explainer_fsd.y_prime
    y_wasserstein_fsd = explainer_fsd.wd.distance(y_s, y_t, delta)[0].item()

    # Record metrics
    run_time = round((time.time() - start_time)/60, 2)
    metrics_fsd_list.append({
        "seed": current_seed,
        "X_SWD_fsd": round(x_wasserstein_fsd, 6),
        "Y_WD_fsd": round(y_wasserstein_fsd, 6),
        "run_time_min": run_time
    })

    print(f"X Sliced Wasserstein Distance (SWD) with FSD constraint: {x_wasserstein_fsd:.6f}")
    print(f"Y Wasserstein Distance (WD) with FSD constraint: {y_wasserstein_fsd:.6f}")

# Metrics summary
df_metrics_fsd = pd.DataFrame(metrics_fsd_list)
ws_metrics = wb_fsd.create_sheet(title="FSD_Metrics_Summary")
for r in dataframe_to_rows(df_metrics_fsd, index=False, header=True):
    ws_metrics.append(r)
wb_fsd.save("market_FSDConstraint_NumDealsPurchases_0.004.xlsx")